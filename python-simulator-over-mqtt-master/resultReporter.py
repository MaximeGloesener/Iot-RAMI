import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from constants import MqttAppConstants

class ExcelGeneratorConstant:
    SENDER_TIME = "sender_time"
    SENDER_VALUE = "sender_value"
    MQTT_TRANSIT = "mqtt_transit"
    RECEIVER_TIME = "receiver_time"
    RECEIVER_VALUE = "receiver_value"

    @staticmethod
    def get_all_rows():
        return [
            ExcelGeneratorConstant.SENDER_TIME,
            ExcelGeneratorConstant.SENDER_VALUE,
            ExcelGeneratorConstant.MQTT_TRANSIT,
            ExcelGeneratorConstant.RECEIVER_TIME,
            ExcelGeneratorConstant.RECEIVER_VALUE
        ]

    @staticmethod
    def get_col_position_on_excel(rowName):
        try:
            # Return the column position (0-based index)
            return ExcelGeneratorConstant.get_all_rows().index(rowName)
        except ValueError:
            # If the column name does not exist, raise an exception
            raise ValueError(f"The column name '{rowName}' does not exist in the defined columns.")


class ResultReporter:
    excel_row_names = ExcelGeneratorConstant.get_all_rows()

    @staticmethod
    def generate_excel(broker_name, mode, received_values):
        # Create a folder to store the results if it doesn't exist
        results_folder = './results'
        if not os.path.exists(results_folder):
            os.makedirs(results_folder)

        # Convert the data into a DataFrame
        df = None
        start_col = None

        if mode == MqttAppConstants.MODE_SENSOR:
            # HERE, only sender_time and sender_value belongs to df
            df = pd.DataFrame(received_values, columns=[ExcelGeneratorConstant.SENDER_TIME, ExcelGeneratorConstant.SENDER_VALUE])
            start_col = ExcelGeneratorConstant.get_col_position_on_excel(ExcelGeneratorConstant.SENDER_TIME)
        elif mode == MqttAppConstants.MODE_SERVER:
            # HERE, only receiver_time and receiver_value belongs to df
            df = pd.DataFrame(received_values, columns=[ExcelGeneratorConstant.RECEIVER_TIME, ExcelGeneratorConstant.RECEIVER_VALUE])
            start_col = ExcelGeneratorConstant.get_col_position_on_excel(ExcelGeneratorConstant.RECEIVER_TIME)

        # Path to the Excel file
        excel_file_path = os.path.join(results_folder, f"{broker_name}.xlsx")

        if os.path.exists(excel_file_path) and mode == MqttAppConstants.MODE_SERVER:
            # If the file exists and the mode is SERVOR, do not overwrite but append data
            book = load_workbook(excel_file_path)
            with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, index=False, sheet_name='Data', startrow=1, startcol=start_col, header=False)            
            ResultReporter.calculate_deltas(broker_name, df, excel_file_path)

        else:
            # Otherwise, create a new file or overwrite the existing one if mode is SENSOR
            with pd.ExcelWriter(excel_file_path, engine='xlsxwriter') as writer:
                # Add time and value
                df.to_excel(writer, index=False, sheet_name='Data', startrow=1, startcol=start_col, header=False)
                # Display the colmun name
                workbook = writer.book
                worksheet = writer.sheets['Data']
                for col_num, value in enumerate(ResultReporter.excel_row_names):
                    worksheet.write(0, col_num, value)
    

    @staticmethod
    def calculate_deltas(broker_name, df, excel_file_path):
        # As, mentionned earlier, here df only have acces to receiver_time and receiver_value
        # However, in order to calculate the delta, i need to get sender_time => let's read it from the file
        book = load_workbook(excel_file_path)
        sheet = book['Data']

        sender_time_col = ExcelGeneratorConstant.get_col_position_on_excel(ExcelGeneratorConstant.SENDER_TIME)+1
        sender_times = [sheet.cell(row=i, column=sender_time_col).value for i in range(2, sheet.max_row + 1)]

        # Convert to datetime
        sender_times = pd.to_datetime(sender_times, format='%H:%M:%S.%f')
        receiver_times = pd.to_datetime(df[ExcelGeneratorConstant.RECEIVER_TIME], format='%H:%M:%S.%f')

        # Calculate the time differences and add the mqtt_transit column
        mqtt_transit_times = (receiver_times - sender_times).dt.total_seconds()
        mqtt_transit_col = ExcelGeneratorConstant.get_col_position_on_excel(ExcelGeneratorConstant.MQTT_TRANSIT)+1

        for i, mqtt_transit in enumerate(mqtt_transit_times, start=2):
            sheet.cell(row=i, column=mqtt_transit_col, value=mqtt_transit)

        book.save(excel_file_path)
        ResultReporter.add_chart_to_excel(broker_name, df, excel_file_path)


    @staticmethod
    def add_chart_to_excel(broker_name, df, excel_file_path):
        # Load the existing file
        book = load_workbook(excel_file_path)
        sheet = book['Data']
        
        # Create a scatter chart
        chart = ScatterChart()
        chart.title = "Transmission time of the nth data sent through the broker {}".format(broker_name)
        chart.x_axis.title = 'nth data sending'
        chart.y_axis.title = 'transmission time (seconds)'

        # Configure the chart series from the DataFrame data
        sender_time_col = ExcelGeneratorConstant.get_col_position_on_excel(ExcelGeneratorConstant.SENDER_TIME) + 1
        mqtt_transit_col = ExcelGeneratorConstant.get_col_position_on_excel(ExcelGeneratorConstant.MQTT_TRANSIT) + 1

        x_values = Reference(sheet, min_col=sender_time_col, min_row=2, max_row=sheet.max_row)
        y_values = Reference(sheet, min_col=mqtt_transit_col, min_row=2, max_row=sheet.max_row)

        series = Series(y_values, x_values, title_from_data=False)
        chart.series.append(series)

        chart.width = 30  # Largeur en pouces
        chart.height = 20  # Hauteur en pouces

        # Add the chart to the sheet
        sheet.add_chart(chart, "G6")

        # Sauvegarder les modifications
        book.save(excel_file_path)
        print(f"Le graphique a été ajouté à {excel_file_path}")